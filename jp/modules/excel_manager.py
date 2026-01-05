from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sqlite3
from .db_manager import DB_PATH

class ExcelManager:
    """엑셀 파일 생성 및 관리를 위한 클래스"""
    
    @staticmethod
    def create_monthly_report(selected_month, transactions):
        """월별 거래명세서 통합 엑셀 파일 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{selected_month}월 거래명세서"
        
        # 스타일 설정
        styles = ExcelManager._get_styles()
        
        # 열 너비 설정
        column_widths = {
            'A': 8,   # 월
            'B': 8,   # 일
            'C': 30,  # 품목
            'D': 15,  # 규격
            'E': 10,  # 수량
            'F': 15,  # 단가
            'G': 15,  # 금액
            'H': 15,  # 미수금
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 공통 헤더 작성
        headers = ['월', '일', '품목', '규격', '수량', '단가', '금액', '미수금']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = styles['header_font']
            cell.alignment = styles['center_align']
            cell.border = styles['header_border']
            cell.fill = styles['header_fill']
        
        # 행 높이 설정
        ws.row_dimensions[1].height = 25  # 헤더 행 높이
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        try:
            current_row = 2  # 헤더 다음 행부터 시작
            total_rows = []  # 각 거래의 합계 행 번호를 저장
            total_balance = 0  # 전체 미수금 누적
            
            for transaction_id, customer_name in transactions:
                final_row = ExcelManager._write_transaction(ws, c, transaction_id, styles, current_row, total_balance)
                total_rows.append(final_row - 2)
                
                # 해당 거래의 총액을 미수금에 추가
                c.execute('''
                    SELECT SUM(supply_price)
                    FROM transaction_items
                    WHERE transaction_id = ?
                ''', (transaction_id,))
                transaction_total = c.fetchone()[0] or 0
                total_balance += transaction_total
                
                current_row = final_row
            
            # 최종 합계 행 추가
            current_row += 1  # 한 줄 띄우기
            
            # 최종 합계 레이블과 금액
            ws.merge_cells(f'A{current_row}:G{current_row}')
            label_cell = ws.cell(row=current_row, column=1)
            label_cell.value = "월 합계 금액"
            label_cell.font = Font(name='맑은 고딕', size=14, bold=True)
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # 각 거래의 합계 금액만 더하는 수식 생성
            sum_formula = f"=SUM(" + ",".join([f"G{row}" for row in total_rows]) + ")"
            
            amount_cell = ws.cell(row=current_row, column=8)
            amount_cell.value = sum_formula
            amount_cell.font = Font(name='맑은 고딕', size=14, bold=True)
            amount_cell.alignment = styles['center_align']
            amount_cell.number_format = '#,##0'
            
            # 최종 합계 행 스타일
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='double'),
                    bottom=Side(style='double')
                )
                cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
            
            # 행 높이 설정
            ws.row_dimensions[current_row].height = 30
            
            # 인쇄 영역 설정
            ws.print_area = f'A1:H{current_row}'
            
            # 인쇄 설정
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT  # 세로 방향
            ws.page_setup.paperSize = ws.PAPERSIZE_A4  # A4 용지
            ws.page_setup.fitToPage = True  # 용지에 맞춤
            ws.page_setup.fitToHeight = False  # 높이는 자동
            ws.page_setup.fitToWidth = 1  # 너비는 1페이지로
            
            return wb
            
        finally:
            conn.close()

    @staticmethod
    def _write_transaction(ws, cursor, transaction_id, styles, start_row, starting_balance):
        """개별 거래명세서 작성"""
        # 메모 정보 조회
        cursor.execute('''
            SELECT memo
            FROM transactions
            WHERE id = ?
        ''', (transaction_id,))
        memo = cursor.fetchone()[0]
        
        # 메모 작성
        if memo:
            # 병합하기 전에 모든 셀에 테두리와 배경색 적용
            for col in range(1, 8):  # A to G columns
                cell = ws.cell(row=start_row, column=col)
                cell.border = Border(
                    left=Side(style='thin') if col > 1 else Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.fill = styles['memo_fill']
            
            # 셀 병합 및 메모 입력
            ws.merge_cells(f'A{start_row}:H{start_row}')
            cell = ws.cell(row=start_row, column=1)
            cell.value = memo
            cell.font = styles['title_font']
            cell.alignment = styles['center_align']
            
            ws.row_dimensions[start_row].height = 20
            start_row += 1

        # 항목 데이터 조회 및 입력
        cursor.execute('''
            SELECT date, product, quantity, unit_price, supply_price, width, height
            FROM transaction_items
            WHERE transaction_id = ?
            ORDER BY date
        ''', (transaction_id,))
        items = cursor.fetchall()
        
        final_row = ExcelManager._write_items(ws, items, styles, start_row, starting_balance)
        return final_row + 1

    @staticmethod
    def _write_items(ws, items, styles, start_row, starting_balance):
        """항목 데이터 작성"""
        running_balance = starting_balance  # 이전 거래들의 누적 미수금으로 시작
        
        for idx, item in enumerate(items):
            row = start_row + idx
            date = item[0].split('-')  # YYYY-MM-DD
            
            # 금액을 running_balance에 더함
            running_balance += item[4]  # supply_price를 누적
            
            # 데이터 입력
            data = [
                int(date[1]),  # 월
                int(date[2]),  # 일
                item[1],       # 품목
                f"{item[5]} X {item[6]}" if item[5] and item[6] else "",  # 규격
                item[2],       # 수량
                item[3],       # 단가
                item[4],       # 금액
                running_balance,  # 누적 미수금
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = styles['normal_font']
                cell.alignment = styles['center_align']
                cell.border = styles['data_border']
                
                if col in [5, 6, 7, 8]:  # 수량, 단가, 금액, 미수금
                    cell.number_format = '#,##0'
            
            ws.row_dimensions[row].height = 18  # 데이터 행 높이
        
        # 합계 행
        total_row = start_row + len(items)
        
        # 병합된 셀의 모든 셀에 테두리 적용
        for col in range(1, 5):  # A to D columns
            cell = ws.cell(row=total_row, column=col)
            cell.border = Border(
                left=Side(style='thin') if col > 1 else Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='double'),
                bottom=Side(style='double')
            )
            cell.fill = styles['total_fill']
        
        # A열 셀에 "합계" 입력
        ws.cell(row=total_row, column=1, value="합계")
        ws.merge_cells(f'A{total_row}:D{total_row}')
        
        # 수량, 단가, 금액, 미수금 합계 및 스타일 적용
        for col, formula in [
            (5, f"=SUM(E{start_row}:E{total_row-1})"),  # 수량
            (6, f"=SUM(F{start_row}:F{total_row-1})"),  # 단가
            (7, f"=SUM(G{start_row}:G{total_row-1})"),  # 금액
            (8, f"=H{total_row-1}")  # 미수금은 마지막 항목의 누적 미수금 값 사용
        ]:
            cell = ws.cell(row=total_row, column=col)
            cell.value = formula
            cell.font = styles['title_font']
            cell.alignment = styles['center_align']
            cell.border = styles['total_border']
            cell.fill = styles['total_fill']
            cell.number_format = '#,##0'
        
        # 합계 행의 첫 번째 셀 스타일
        first_cell = ws.cell(row=total_row, column=1)
        first_cell.font = styles['title_font']
        first_cell.alignment = styles['center_align']
        
        ws.row_dimensions[total_row].height = 20  # 합계 행 높이
        
        return total_row + 1

    @staticmethod
    def _get_styles():
        """스타일 설정 반환"""
        header_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        
        data_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        total_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='double'),
            bottom=Side(style='double')
        )
        
        return {
            'header_font': Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF'),
            'title_font': Font(name='맑은 고딕', size=11, bold=True),
            'normal_font': Font(name='맑은 고딕', size=10),
            'final_font': Font(name='맑은 고딕', size=14, bold=True),
            'center_align': Alignment(horizontal='center', vertical='center'),
            'header_border': header_border,
            'data_border': data_border,
            'total_border': total_border,
            'header_fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'memo_fill': PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid'),
            'total_fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
            'final_fill': PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        } 
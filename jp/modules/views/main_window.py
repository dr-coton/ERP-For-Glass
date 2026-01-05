import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from modules.views.customer_manager import CustomerManager
from modules.views.product_manager import ProductManager
from modules.views.transaction_manager import TransactionManager
from modules.views.transaction_viewer import TransactionViewer
import sqlite3
from tkinter import messagebox
from modules.db_manager import DB_PATH
from openpyxl import Workbook, load_workbook
from tkinter import filedialog
import os
import platform
import subprocess
from ..excel_manager import ExcelManager

class MainWindow:
    def __init__(self, root):
        self.root = root
        
        # 메인 컨테이너 생성
        self.main_container = ttk.Frame(root, padding="10")
        self.main_container.pack(fill=BOTH, expand=True)
        
        # 상단 타이틀 프레임
        self.title_frame = ttk.Frame(self.main_container)
        self.title_frame.pack(fill=X, pady=(0, 10))
        
        self.title_label = ttk.Label(
            self.title_frame,
            text="거래명세서 관리 시스템",
            font=("Helvetica", 16, "bold")
        )
        self.title_label.pack(side=LEFT)
        
        # 툴바 프레임
        self.toolbar = ttk.Frame(self.main_container)
        self.toolbar.pack(fill=X, pady=(0, 10))
        
        # 왼쪽 버튼 그룹
        self.left_buttons = ttk.Frame(self.toolbar)
        self.left_buttons.pack(side=LEFT)
        
        # 버튼들에 아이콘과 스타일 적용
        self.add_btn = ttk.Button(
            self.left_buttons,
            text="새 거래명세서",
            style="primary.TButton",
            padding=(10, 5)
        )
        self.add_btn.pack(side=LEFT, padx=(0, 5))
        
        self.customer_btn = ttk.Button(
            self.left_buttons,
            text="거래처 관리",
            style="secondary.TButton",
            padding=(10, 5)
        )
        self.customer_btn.pack(side=LEFT, padx=(0, 5))
        
        self.product_btn = ttk.Button(
            self.left_buttons,
            text="상품 관리",
            style="secondary.TButton",
            padding=(10, 5)
        )
        self.product_btn.pack(side=LEFT, padx=(0, 5))
        
        # 검색 프레임 (오른쪽)
        self.search_frame = ttk.Frame(self.toolbar)
        self.search_frame.pack(side=RIGHT)
        
        self.search_entry = ttk.Entry(self.search_frame, width=30)
        self.search_entry.pack(side=LEFT, padx=(0, 5))
        
        self.search_btn = ttk.Button(
            self.search_frame,
            text="검색",
            style="info.TButton"
        )
        self.search_btn.pack(side=LEFT)
        
        # 메인 콘텐츠 영역
        self.content_frame = ttk.Frame(self.main_container)
        self.content_frame.pack(fill=BOTH, expand=True)
        
        # 거래명세서 목록 트리뷰
        columns = ("id", "created_at", "customer_name", "total_amount", "memo")
        self.tree = ttk.Treeview(
            self.content_frame,
            columns=columns,
            show="headings",
            selectmode="extended",
            style="Custom.Treeview"
        )
        
        # 트리뷰 스타일 설정
        style = ttk.Style()
        
        # Treeview 스타일 설정
        style.configure(
            "Custom.Treeview", 
            rowheight=25,
            font=('Helvetica', 10)
        )
        
        # Treeview Heading 스타일 설정
        style.configure(
            "Custom.Treeview.Heading",
            font=('Helvetica', 10, 'bold')
        )
        
        # Heading 배경색 설정
        style.map(
            "Custom.Treeview.Heading",
            background=[('', '#E8E8E8')]  # 모든 상태에서 회색 배경 유지
        )
        
        # 선택된 항목 색상 설정 (선택 사항)
        style.map(
            "Custom.Treeview",
            background=[('selected', '#0078D7')],
            foreground=[('selected', 'white')]
        )
        
        # 컬럼 설정
        self.tree.heading("id", text="번호", anchor="center")
        self.tree.heading("created_at", text="작성일", anchor="center")
        self.tree.heading("customer_name", text="거래처명", anchor="center")
        self.tree.heading("total_amount", text="총 금액", anchor="center")
        self.tree.heading("memo", text="메모", anchor="w")
        
        # 컬럼 너비 및 정렬 설정
        self.tree.column("id", width=80, anchor="center")
        self.tree.column("created_at", width=150, anchor="center")
        self.tree.column("customer_name", width=200, anchor="center")
        self.tree.column("total_amount", width=150, anchor="center")
        self.tree.column("memo", width=300, anchor="w")
        
        # 스크롤바 추가
        scrollbar = ttk.Scrollbar(self.content_frame, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # 트리뷰와 스크롤바 배치
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        # 트리뷰 더블클릭 이벤트 바인딩 수정
        self.tree.bind('<Double-1>', self.edit_transaction)
        
        # 버튼 프레임 추가 (트리뷰 아래)
        self.action_buttons = ttk.Frame(self.main_container)
        self.action_buttons.pack(fill=X, pady=(10, 0))
        
        # 작업 버튼들 - 초기에는 비활성화 상태로 생성
        self.view_btn = ttk.Button(
            self.action_buttons,
            text="상세 보기",
            style="info.TButton",
            command=self.view_transaction_details,
            state="disabled"
        )
        self.view_btn.pack(side=LEFT, padx=(0, 5))
        
        self.download_btn = ttk.Button(
            self.action_buttons,
            text="다운로드",
            style="info.TButton",
            command=self.download_transaction,
            state="disabled"
        )
        self.download_btn.pack(side=LEFT, padx=(0, 5))
        
        # 월별 다운로드 버튼 추가
        self.monthly_download_btn = ttk.Button(
            self.action_buttons,
            text="월별 다운로드",
            style="info.TButton",
            command=self.download_monthly_transactions
        )
        self.monthly_download_btn.pack(side=LEFT, padx=(0, 5))
        
        self.edit_btn = ttk.Button(
            self.action_buttons,
            text="수정하기",
            style="warning.TButton",
            command=self.edit_transaction,
            state="disabled"
        )
        self.edit_btn.pack(side=LEFT, padx=(0, 5))
        
        self.delete_btn = ttk.Button(
            self.action_buttons,
            text="삭제하기",
            style="danger.TButton",
            command=self.delete_transaction,
            state="disabled"
        )
        self.delete_btn.pack(side=LEFT)
        
        # 초기 데이터 로드
        self.load_transactions()
        
        # 거래처 관리 버튼에 기능 연결
        self.customer_btn.config(command=self.open_customer_manager)
        
        # 상품 관리 버튼에 기능 연결
        self.product_btn.config(command=self.open_product_manager)
        
        # 새 거래명세서 버튼에 기능 연결
        self.add_btn.config(command=self.open_transaction_manager)
        
        # 검색 버튼에 기능 연결
        self.search_btn.config(command=self.search_transactions)
        
        # 트리뷰 선택 이벤트 바인딩
        self.tree.bind('<<TreeviewSelect>>', self.on_select)
        
    def open_customer_manager(self):
        # 거래처 관리 창 열기
        customer_manager = CustomerManager(self.root) 
    
    def open_product_manager(self):
        product_manager = ProductManager(self.root) 
    
    def open_transaction_manager(self):
        transaction_manager = TransactionManager(self.root, self) 
    
    def load_transactions(self):
        """거래명세서 목록 로드"""
        # 기존 항목 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute('''
                SELECT 
                    t.id,
                    COALESCE(
                        (SELECT date 
                         FROM transaction_items 
                         WHERE transaction_id = t.id 
                         ORDER BY date ASC 
                         LIMIT 1),
                        strftime('%Y-%m-%d', t.created_at)
                    ) as display_date,
                    t.customer_name,
                    t.total_amount,
                    t.memo
                FROM transactions t
                ORDER BY display_date DESC
            ''')
            transactions = c.fetchall()
            
            # 트리뷰에 데이터 추가
            for transaction in transactions:
                formatted_amount = f"{transaction[3]:,}원"
                values = (
                    transaction[0],
                    transaction[1],  # 항목의 첫 번째 날짜 또는 created_at
                    transaction[2],
                    formatted_amount,
                    transaction[4] or ""
                )
                self.tree.insert('', 'end', values=values)
                
        except Exception as e:
            messagebox.showerror("오류", f"거래명세서 목록 로드 중 오류 발생: {str(e)}")
        finally:
            if conn:
                conn.close()
    
    def search_transactions(self):
        """거래명세서 검색"""
        search_text = self.search_entry.get().strip()
        
        # 기존 항목 삭제
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute('''
                SELECT id, customer_name, total_amount, created_at, memo 
                FROM transactions 
                WHERE customer_name LIKE ? OR memo LIKE ?
                ORDER BY created_at DESC
            ''', (f'%{search_text}%', f'%{search_text}%'))
            transactions = c.fetchall()
            
            # 트리뷰에 데이터 추가
            for transaction in transactions:
                formatted_amount = f"{transaction[2]:,}원"
                values = (
                    transaction[0],
                    transaction[1],
                    formatted_amount,
                    transaction[3],
                    transaction[4] or ""
                )
                self.tree.insert('', 'end', values=values)
                
        except Exception as e:
            messagebox.showerror("오류", f"거래명세서 검색 중 오류 발생: {str(e)}")
        finally:
            if conn:
                conn.close() 
    
    def view_transaction_details(self, event=None):
        """거래명세서 상세 보기"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("경고", "거래명세서를 선택하세요.")
            return
        
        transaction_id = self.tree.item(selected[0])['values'][0]
        TransactionViewer(self.root, transaction_id)
    
    def download_transaction(self):
        """거래명세서 엑셀 다운로드"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("경고", "다운로드할 거래명세서를 선택하세요.")
            return
        
        transaction_id = self.tree.item(selected[0])['values'][0]
        
        try:
            # 템플릿 파일 로드
            template_path = os.path.join("templates", "transaction_template.xlsx")
            if not os.path.exists(template_path):
                messagebox.showerror("오류", "엑셀 템플릿 파일을 찾을 수 없습니다.")
                return
            
            wb = load_workbook(template_path)
            ws = wb.active
            
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            
            # 거래명세서 기본 정보 조회
            c.execute('''
                SELECT 
                    t.id,
                    t.customer_name,
                    t.total_amount,
                    t.memo,
                    t.created_at,
                    c.address,
                    c.business_id,
                    c.representative,
                    c.phone
                FROM transactions t
                LEFT JOIN customers c ON t.customer_name = c.company_name
                WHERE t.id = ?
            ''', (transaction_id,))
            transaction = c.fetchone()
            
            if not transaction:
                messagebox.showerror("오류", "거래명세서 정보를 찾을 수 없습니다.")
                return
            
            # 거래명세서 항목 조회
            c.execute('''
                SELECT date, product, quantity, unit_price, supply_price, width, height
                FROM transaction_items
                WHERE transaction_id = ?
                ORDER BY date
            ''', (transaction_id,))
            items = c.fetchall()
            
            # 기본 정보 입력 (셀 위치는 템플릿에 맞게 조정)
            ws['H4'] = transaction[1] or ""  # 거래처명
            ws['H5'] = transaction[5] or ""  # 주소
            ws['H6'] = transaction[8] or ""  # 전화번호
            ws['H7'] = f"{transaction[2]:,}원"  # 총 금액을 콤마가 포함된 형식으로 표시
            
            # 항목 데이터 입력 (9행부터 시작)
            start_row = 9
            for idx, item in enumerate(items):
                row = start_row + idx
                date = item[0].split('-')  # YYYY-MM-DD 형식에서 월과 일 추출
                
                ws.cell(row=row, column=2, value=int(date[1]))   # 월 (B열)
                ws.cell(row=row, column=3, value=int(date[2]))   # 일 (C열)
                ws.cell(row=row, column=4, value=item[1])        # 품목명 (D열)
                
                # width와 height가 있는 경우 사이즈 정보 표시
                size_info = ""
                if item[5] and item[6]:  # width와 height가 있는 경우
                    size_info = f"{item[5]} X {item[6]}"
                ws.cell(row=row, column=11, value=size_info)     # 사이즈 (K열)
                
                ws.cell(row=row, column=17, value=item[2])       # 수량 (Q열)
                ws.cell(row=row, column=21, value=item[3])       # 단가 (U열)
                ws.cell(row=row, column=27, value=item[4])       # 공급가액 (AA열)
                
                # 숫자 셀 서식 적용
                for col in [17, 21, 27]:  # 수량, 단가, 공급가액에 천단위 콤마 적용
                    ws.cell(row=row, column=col).number_format = '#,##0'
            
            # 항목의 첫 번째 날짜 조회
            c.execute('''
                SELECT COALESCE(
                    (SELECT date 
                     FROM transaction_items 
                     WHERE transaction_id = ? 
                     ORDER BY date ASC 
                     LIMIT 1),
                    strftime('%Y-%m-%d', t.created_at)
                ) as display_date
                FROM transactions t
                WHERE t.id = ?
            ''', (transaction_id, transaction_id))
            display_date = c.fetchone()[0]
            
            # 파일 저장
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{transaction_id}_{display_date}_{transaction[1]}.xlsx"  # 관리번호_날짜_거래처명
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("성공", "거래명세서가 엑셀 파일로 저장되었습니다.")
                
                # 저장된 파일 열기
                if messagebox.askyesno("확인", "저장된 파일을 여어보시겠습니까?"):
                    try:
                        if platform.system() == 'Darwin':       # macOS
                            subprocess.call(('open', file_path))
                        elif platform.system() == 'Windows':    # Windows
                            os.startfile(file_path)
                        else:                                   # linux variants
                            subprocess.call(('xdg-open', file_path))
                    except Exception as e:
                        messagebox.showerror("오류", f"파일을 여는 중 오류가 발생했습니다: {str(e)}")
        
        except Exception as e:
            messagebox.showerror("오류", f"엑셀 파일 생성 중 오류가 발생했습니다: {str(e)}")
        finally:
            if conn:
                conn.close()
    
    def download_monthly_transactions(self):
        """월별 거래명세서 통합 엑셀 다운로드"""
        try:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()

            # 선택된 월의 모든 거래명세서 조회
            c.execute('''
                SELECT DISTINCT strftime('%Y-%m', 
                    COALESCE(
                        (SELECT date FROM transaction_items 
                         WHERE transaction_id = t.id 
                         ORDER BY date ASC LIMIT 1),
                        t.created_at
                    )
                ) as month
                FROM transactions t
                ORDER BY month DESC
            ''')
            months = c.fetchall()

            if not months:
                messagebox.showinfo("알림", "다운로드할 거래명세서가 없습니다.")
                return

            # 월 선택 다이얼로그 표시 및 거래명세서 조회 로직
            selected_month = self._show_month_selection_dialog(months)
            if not selected_month:
                return

            transactions = self._get_monthly_transactions(c, selected_month)
            if not transactions:
                messagebox.showinfo("알림", f"{selected_month} 월에 해당하는 거래명세서가 없습니다.")
                return

            # 엑셀 파일 생성
            wb = ExcelManager.create_monthly_report(selected_month, transactions)

            # 파일 저장
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{selected_month}_거래명세서.xlsx"
            )
            
            if file_path:
                wb.save(file_path)
                messagebox.showinfo("성공", f"{selected_month} 월 거래명세서가 저장되었습니다.")
                self._open_file(file_path)

        except Exception as e:
            messagebox.showerror("오류", f"월별 다운로드 중 오류가 발생했습니다: {str(e)}")
        finally:
            if conn:
                conn.close()

    def _show_month_selection_dialog(self, months):
        """월 선택 다이얼로그 표시"""
        month_dialog = ttk.Toplevel(self.root)
        month_dialog.title("월 선택")
        month_dialog.geometry("300x400")

        month_listbox = ttk.Treeview(
            month_dialog,
            columns=("month",),
            show="headings",
            selectmode="browse"
        )
        month_listbox.heading("month", text="년월")
        month_listbox.pack(fill=BOTH, expand=True, padx=10, pady=10)

        for month in months:
            month_listbox.insert('', 'end', values=(month[0],))

        selected_month = [None]

        def on_select():
            selection = month_listbox.selection()
            if selection:
                selected_month[0] = month_listbox.item(selection[0])['values'][0]
                month_dialog.destroy()

        ttk.Button(
            month_dialog,
            text="선택",
            command=on_select
        ).pack(pady=10)

        month_dialog.transient(self.root)
        month_dialog.grab_set()
        self.root.wait_window(month_dialog)

        return selected_month[0]

    def _get_monthly_transactions(self, cursor, selected_month):
        """선택된 월의 거래명세서 목록 조회"""
        cursor.execute('''
            SELECT t.id, t.customer_name
            FROM transactions t
            WHERE strftime('%Y-%m', 
                COALESCE(
                    (SELECT date FROM transaction_items 
                     WHERE transaction_id = t.id 
                     ORDER BY date ASC LIMIT 1),
                    t.created_at
                )
            ) = ?
            AND t.customer_name = '왕성글래스'
            ORDER BY t.customer_name
        ''', (selected_month,))
        return cursor.fetchall()

    def _open_file(self, file_path):
        """파일 열기"""
        if messagebox.askyesno("확인", "저장된 파일을 열어보시겠습니까?"):
            try:
                if platform.system() == 'Darwin':       # macOS
                    subprocess.call(('open', file_path))
                elif platform.system() == 'Windows':    # Windows
                    os.startfile(file_path)
                else:                                   # linux variants
                    subprocess.call(('xdg-open', file_path))
            except Exception as e:
                messagebox.showerror("오류", f"파일을 여는 중 오류가 발생했습니다: {str(e)}")
    
    def edit_transaction(self, event=None):
        """거래명세서 수정"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("경고", "수정할 거래명세서를 선택하세요.")
            return
        
        transaction_id = self.tree.item(selected[0])['values'][0]
        transaction_manager = TransactionManager(self.root, self, transaction_id)
    
    def delete_transaction(self):
        """거래명세서 삭제"""
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("경고", "삭제할 거래명세서를 선택하세요.")
            return
        
        if not messagebox.askyesno("확인", "선택한 거래명세서를 삭제하시겠습니까?"):
            return
        
        try:
            transaction_id = self.tree.item(selected[0])['values'][0]
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            
            # 거래명세서 항목 삭제
            c.execute('DELETE FROM transaction_items WHERE transaction_id=?', (transaction_id,))
            # 거래명세서 삭제
            c.execute('DELETE FROM transactions WHERE id=?', (transaction_id,))
            
            conn.commit()
            conn.close()
            
            self.load_transactions()
            messagebox.showinfo("성공", "거래명세서가 삭제되었습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"거래명세서 삭제 중 오류가 발생했습니다: {str(e)}") 
    
    def on_select(self, event=None):
        """트리뷰에서 항목이 선택되었을 때의 처리"""
        selected = self.tree.selection()
        
        # 버튼들의 상태를 선택 여부에 따라 설정
        state = "normal" if selected else "disabled"
        self.view_btn.configure(state=state)
        self.download_btn.configure(state=state)
        self.edit_btn.configure(state=state)
        self.delete_btn.configure(state=state) 
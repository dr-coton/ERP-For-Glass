from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

import sqlite3
from openpyxl import load_workbook

from modules.db_manager import DB_PATH
from modules.excel_manager import ExcelManager

_TEMPLATE_PATH = Path("templates") / "transaction_template.xlsx"


class TransactionNotFoundError(RuntimeError):
    """Raised when the requested transaction does not exist."""


def _sanitize_filename(text: str) -> str:
    cleaned = "_".join(text.strip().split()) or "transaction"
    return cleaned.replace("/", "-")


def build_transaction_workbook(transaction_id: int) -> tuple[BytesIO, str]:
    if not _TEMPLATE_PATH.exists():
        raise FileNotFoundError("엑셀 템플릿이 존재하지 않습니다. templates/transaction_template.xlsx 파일을 확인하세요.")

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT 
                t.id,
                t.customer_name,
                t.total_amount,
                t.memo,
                t.created_at,
                c.address,
                c.business_id,
                c.representative,
                c.phone
            FROM transactions t
            LEFT JOIN customers c ON t.customer_name = c.company_name
            WHERE t.id = ?
            """,
            (transaction_id,),
        )
        transaction = cursor.fetchone()
        if transaction is None:
            raise TransactionNotFoundError(f"거래명세서 {transaction_id} 번을 찾을 수 없습니다.")

        cursor.execute(
            """
            SELECT date, product, quantity, unit_price, supply_price, width, height
            FROM transaction_items
            WHERE transaction_id = ?
            ORDER BY date
            """,
            (transaction_id,),
        )
        items: Iterable[sqlite3.Row] = cursor.fetchall()
    finally:
        conn.close()

    workbook = load_workbook(_TEMPLATE_PATH)
    sheet = workbook.active

    sheet["H4"] = transaction["customer_name"] or ""
    sheet["H5"] = transaction["address"] or ""
    sheet["H6"] = transaction["phone"] or ""
    sheet["H7"] = f"{transaction['total_amount']:,}원"

    start_row = 9
    for index, item in enumerate(items):
        current_row = start_row + index
        date_parts = (item["date"] or "0000-00-00").split("-")
        try:
            month_part = int(date_parts[1])
            day_part = int(date_parts[2])
        except (IndexError, ValueError):
            month_part = day_part = 0

        sheet.cell(row=current_row, column=2, value=month_part)
        sheet.cell(row=current_row, column=3, value=day_part)
        sheet.cell(row=current_row, column=4, value=item["product"])

        size_info = ""
        if item["width"] and item["height"]:
            size_info = f"{item['width']} X {item['height']}"
        sheet.cell(row=current_row, column=11, value=size_info)

        sheet.cell(row=current_row, column=17, value=item["quantity"])
        sheet.cell(row=current_row, column=21, value=item["unit_price"])
        sheet.cell(row=current_row, column=27, value=item["supply_price"])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{_sanitize_filename(transaction['customer_name'] or '거래명세서')}_{transaction_id}.xlsx"
    return output, filename


def build_monthly_workbook(year_month: str) -> tuple[BytesIO, str]:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    try:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT t.id, t.customer_name
            FROM transactions t
            WHERE strftime('%Y-%m', 
                COALESCE(
                    (SELECT date FROM transaction_items 
                     WHERE transaction_id = t.id 
                     ORDER BY date ASC LIMIT 1),
                    t.created_at
                )
            ) = ?
            AND t.customer_name = '왕성글래스'
            ORDER BY t.customer_name
            """,
            (year_month,),
        )
        rows = cursor.fetchall()
    finally:
        conn.close()

    if not rows:
        raise TransactionNotFoundError(f'{year_month}에 해당하는 거래명세서를 찾을 수 없습니다.')

    transactions = [(row['id'], row['customer_name']) for row in rows]
    workbook = ExcelManager.create_monthly_report(year_month, transactions)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"{year_month}_거래명세서.xlsx"
    return output, filename

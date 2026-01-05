from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import sqlite3
from datetime import datetime, date
from modules.db_manager import DB_PATH
import os

def generate_customer_statement(customer_name, start_date, end_date=None):
    """특정 고객의 거래명세서를 생성"""
    
    if end_date is None:
        end_date = date.today().strftime('%Y-%m-%d')
    
    # 데이터베이스 연결
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    
    # 거래 데이터 조회
    c.execute('''
        SELECT t.id, t.customer_name, t.total_amount, t.memo, t.created_at
        FROM transactions t
        WHERE t.customer_name = ? 
        AND DATE(t.created_at) >= ? 
        AND DATE(t.created_at) <= ?
        ORDER BY t.created_at
    ''', (customer_name, start_date, end_date))
    
    transactions = c.fetchall()
    
    if not transactions:
        print(f"{customer_name}의 {start_date}부터 {end_date}까지의 거래내역이 없습니다.")
        conn.close()
        return None
    
    # 엑셀 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = f"{customer_name} 거래명세서"
    
    # 스타일 설정
    header_font = Font(name='맑은 고딕', size=12, bold=True)
    title_font = Font(name='맑은 고딕', size=14, bold=True)
    normal_font = Font(name='맑은 고딕', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    
    # 제목 작성
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = f"{customer_name} 거래명세서"
    title_cell.font = Font(name='맑은 고딕', size=16, bold=True)
    title_cell.alignment = center_align
    
    # 기간 표시
    ws.merge_cells('A2:I2')
    period_cell = ws['A2']
    period_cell.value = f"기간: {start_date} ~ {end_date}"
    period_cell.font = Font(name='맑은 고딕', size=12)
    period_cell.alignment = center_align
    
    # 열 너비 설정
    column_widths = {
        'A': 12,  # 날짜
        'B': 15,  # 품목
        'C': 12,  # 단가종류
        'D': 10,  # 가로
        'E': 10,  # 세로
        'F': 8,   # 수량
        'G': 12,  # 단가
        'H': 15,  # 공급가액
        'I': 20,  # 메모
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # 헤더 작성
    headers = ['날짜', '품목', '단가종류', '가로', '세로', '수량', '단가', '공급가액', '메모']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.alignment = center_align
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    current_row = 5
    total_amount = 0
    
    # 각 거래별 항목 작성
    for transaction in transactions:
        transaction_id = transaction[0]
        memo = transaction[3] or ""
        
        # 거래 항목 조회
        c.execute('''
            SELECT date, product, price_type, width, height, 
                   quantity, unit_price, supply_price
            FROM transaction_items
            WHERE transaction_id = ?
            ORDER BY date
        ''', (transaction_id,))
        
        items = c.fetchall()
        
        for item in items:
            # 데이터 입력
            ws.cell(row=current_row, column=1).value = item[0]  # 날짜
            ws.cell(row=current_row, column=2).value = item[1]  # 품목
            ws.cell(row=current_row, column=3).value = item[2]  # 단가종류
            ws.cell(row=current_row, column=4).value = item[3]  # 가로
            ws.cell(row=current_row, column=5).value = item[4]  # 세로
            ws.cell(row=current_row, column=6).value = item[5]  # 수량
            ws.cell(row=current_row, column=7).value = f"{item[6]:,}"  # 단가
            ws.cell(row=current_row, column=8).value = f"{item[7]:,}"  # 공급가액
            ws.cell(row=current_row, column=9).value = memo  # 메모
            
            # 스타일 적용
            for col in range(1, 10):
                cell = ws.cell(row=current_row, column=col)
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center' if col != 9 else 'left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            total_amount += item[7]  # 공급가액 누적
            current_row += 1
    
    # 합계 행 추가
    current_row += 1
    ws.merge_cells(f'A{current_row}:G{current_row}')
    total_label_cell = ws.cell(row=current_row, column=1)
    total_label_cell.value = "합계"
    total_label_cell.font = title_font
    total_label_cell.alignment = Alignment(horizontal='right', vertical='center')
    
    total_cell = ws.cell(row=current_row, column=8)
    total_cell.value = f"{total_amount:,}"
    total_cell.font = title_font
    total_cell.alignment = center_align
    
    # 합계 행 스타일
    for col in range(1, 10):
        cell = ws.cell(row=current_row, column=col)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='double'),
            bottom=Side(style='double')
        )
        cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    
    conn.close()
    
    # 파일 저장
    filename = f"{customer_name}_거래명세서_{start_date}_{end_date}.xlsx"
    filepath = os.path.join("exports", filename)
    
    # exports 디렉토리가 없으면 생성
    if not os.path.exists("exports"):
        os.makedirs("exports")
    
    wb.save(filepath)
    print(f"거래명세서가 생성되었습니다: {filepath}")
    return filepath

# 조은유리의 6월 17일부터 현재까지 거래명세서 생성
if __name__ == "__main__":
    customer_name = "조은유리"
    start_date = "2024-06-17"
    
    filepath = generate_customer_statement(customer_name, start_date)
    if filepath:
        print(f"\n✅ 거래명세서 생성 완료!")
        print(f"📁 파일 위치: {filepath}")